"""
Prosty storage dla draftów (JSON) + import Excel/CSV.
"""
from __future__ import annotations
import json
import os
from datetime import datetime
from typing import List, Optional
import glob

from app.models import GLBZ2_Form, GIR_Form


DRAFTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "drafts")
os.makedirs(DRAFTS_DIR, exist_ok=True)


# ── Zapis/Odczyt draftów ─────────────────────────────────────────────────────

def save_draft(form_type: str, data: dict, name: str = "") -> str:
    """Zapisuje draft do pliku JSON. Zwraca ścieżkę pliku."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = name.replace(" ", "_")[:40] if name else "draft"
    filename = f"{form_type}_{safe_name}_{ts}.json"
    path = os.path.join(DRAFTS_DIR, filename)
    with open(path, "w", encoding="utf-8") as f:
        json.dump({"form_type": form_type, "saved_at": ts, "data": data}, f, ensure_ascii=False, indent=2)
    return path


def list_drafts(form_type: Optional[str] = None) -> List[dict]:
    """Zwraca listę zapisanych draftów z metadanymi."""
    pattern = os.path.join(DRAFTS_DIR, f"{form_type}_*.json" if form_type else "*.json")
    drafts = []
    for path in sorted(glob.glob(pattern), reverse=True):
        try:
            with open(path, "r", encoding="utf-8") as f:
                meta = json.load(f)
            drafts.append({
                "path": path,
                "filename": os.path.basename(path),
                "form_type": meta.get("form_type", "?"),
                "saved_at": meta.get("saved_at", ""),
            })
        except Exception:
            pass
    return drafts


def load_draft(path: str) -> dict:
    """Wczytuje draft z pliku JSON."""
    with open(path, "r", encoding="utf-8") as f:
        obj = json.load(f)
    return obj.get("data", obj)


# ── Import Excel/CSV dla jednostek D (GLB-Z2) ────────────────────────────────

GLBZ2_EXCEL_COLUMNS = {
    "rodzaj_jednostki": "Rodzaj jednostki (1/2)",
    "pelna_nazwa": "Pełna nazwa",
    "kraj_id": "Kraj nr ID (ISO)",
    "rodzaj_id": "Rodzaj ID (1/2/3/4/8/9)",
    "nr_id": "Numer identyfikacyjny",
    "adres_kraj": "Kraj siedziby (ISO)",
    "adres_miejscowosc": "Miejscowość",
    "adres_kod": "Kod pocztowy",
    "adres_ulica": "Ulica",
    "adres_nr_budynku": "Nr domu",
    "adres_nr_lokalu": "Nr lokalu",
    "adres_inne": "Inne dane adresowe",
    "kod_jurysdykcji": "Kod jurysdykcji (ISO)",
}


def import_jednostki_from_excel(file_bytes: bytes, filename: str) -> tuple[list, list]:
    """
    Importuje listę jednostek D z pliku Excel (.xlsx) lub CSV.
    Zwraca (lista_dict, lista_błędów).
    """
    import io
    errors = []
    records = []

    try:
        if filename.endswith(".csv"):
            import csv
            reader = csv.DictReader(io.StringIO(file_bytes.decode("utf-8-sig", errors="replace")))
            rows = list(reader)
        else:
            import pandas as pd
            df = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
            df = df.fillna("")
            rows = df.to_dict("records")
    except Exception as e:
        return [], [f"Nie można wczytać pliku: {e}"]

    # Mapowanie: przyjazna nazwa kolumny → nazwa pola modelu
    col_map = {v: k for k, v in GLBZ2_EXCEL_COLUMNS.items()}
    # Alternatywnie akceptujemy też nazwy techniczne bezpośrednio
    col_map.update({k: k for k in GLBZ2_EXCEL_COLUMNS})

    for i, row in enumerate(rows, 1):
        mapped = {}
        for col, val in row.items():
            field = col_map.get(col.strip())
            if field:
                mapped[field] = str(val).strip()
        if not mapped.get("pelna_nazwa"):
            errors.append(f"Wiersz {i}: brakuje pełnej nazwy – wiersz pominięty")
            continue
        records.append(mapped)

    return records, errors


GIR_EXCEL_COLUMNS = {
    "kod_jurysdykcji": "Kod jurysdykcji (ISO)",
    "safe_harbour": "Safe Harbour",
    "etr_range": "Przedział ETR",
    "globe_tut": "GloBE TuT",
    "doc_ref_id": "DocRefId",
    "doc_type": "DocType (OECD1/2/3)",
}


def import_summary_from_excel(file_bytes: bytes, filename: str) -> tuple[list, list]:
    """Importuje listę sekcji Summary GIR z Excel/CSV."""
    import io
    errors = []
    records = []

    try:
        if filename.endswith(".csv"):
            import csv
            reader = csv.DictReader(io.StringIO(file_bytes.decode("utf-8-sig", errors="replace")))
            rows = list(reader)
        else:
            import pandas as pd
            df = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
            df = df.fillna("")
            rows = df.to_dict("records")
    except Exception as e:
        return [], [f"Nie można wczytać pliku: {e}"]

    col_map = {v: k for k, v in GIR_EXCEL_COLUMNS.items()}
    col_map.update({k: k for k in GIR_EXCEL_COLUMNS})

    for i, row in enumerate(rows, 1):
        mapped = {}
        for col, val in row.items():
            field = col_map.get(col.strip())
            if field:
                mapped[field] = str(val).strip()
        if not mapped.get("kod_jurysdykcji"):
            errors.append(f"Wiersz {i}: brakuje kodu jurysdykcji – wiersz pominięty")
            continue
        records.append(mapped)

    return records, errors


def generate_excel_template(form_type: str) -> bytes:
    """Generuje plik Excel z nagłówkami szablonu."""
    import io
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active

    if form_type == "GLBZ2":
        ws.title = "Jednostki_D"
        cols = list(GLBZ2_EXCEL_COLUMNS.values())
    else:
        ws.title = "Summary"
        cols = list(GIR_EXCEL_COLUMNS.values())

    for j, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = 25

    # Przykładowy wiersz
    if form_type == "GLBZ2":
        ws.append(["1", "Example Corp Ltd", "DE", "1", "DE123456789",
                   "DE", "Berlin", "10115", "Unter den Linden", "1", "", "", "DE"])
    else:
        ws.append(["DE", "STSH", "C", "", "GS-DOC001", "OECD1"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
